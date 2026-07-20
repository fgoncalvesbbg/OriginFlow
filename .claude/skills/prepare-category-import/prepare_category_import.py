#!/usr/bin/env python3
"""
prepare-category-import
=======================
Convert a raw category export (one row per attribute, one column per SKU) into the two files the
OriginFlow app imports:

  1. <category>_attributes.csv  -> Admin > Categories > Import from CSV   (import FIRST)
  2. <category>_skus.csv        -> SKU Catalog > Bulk upload              (import SECOND)

Input columns expected: "Attribute", "attribute_akeneo", "Where to store" (+ SKU columns whose
header starts with the SKU prefix, default "100"). Only rows whose "Where to store" contains
"Akeneo" are kept. Akeneo codes are trusted from "attribute_akeneo".

Usage:
  python3 prepare_category_import.py INPUT.xlsx --category "Wine Coolers" [--out-dir docs] [--sku-prefix 100]

.xlsx needs openpyxl (`pip install openpyxl`); .csv needs nothing.
"""
import argparse
import csv
import os
import re
import sys

# ── Output vocab the app understands ─────────────────────────────────────────
ATTR_HEADER = ["Attribute", "Type", "Akeneo Code", "Suggested Data Type", "Options / Range", "Notes"]
DEFAULT_GROUP = "1 . Category Specific Attributes"

# Map free-form group/section text -> a label the app's parseAttributeCsv recognizes.
GROUP_ALIASES = {
    "category & segmentation": "Category & Segmentation",
    "segmentation": "Category & Segmentation",
    "listing & data": "Listing & Data",
    "variation": "Listing & Data",
    "category specific": DEFAULT_GROUP,
    "category specific attributes": DEFAULT_GROUP,
    "standard specs": "2. Standard Specs",
    "standard electric specs": "2. Standard Specs",
    "product dimensions": "3. Product Dimensions",
    "dimensions": "3. Product Dimensions",
    "battery": "4. Battery Information",
    "battery information": "4. Battery Information",
    "packaging": "5. Packaging",
}
TITLE_ROW_NAMES = {"title", "name", "product name", "sku title", "label", "produkt", "product_name"}
BOOL_TRUE = {"yes", "y", "true", "1", "x"}
BOOL_FALSE = {"no", "n", "false", "0"}


def norm(s):
    return str(s if s is not None else "").strip().lower()


def slugify(s):
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", str(s).strip().lower())).strip("_")


# ── Loading ──────────────────────────────────────────────────────────────────
def cell_to_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def load_grid(path):
    """Return the sheet as a list of rows (each a list of trimmed strings)."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl  # type: ignore
        except ImportError:
            sys.exit("This is an .xlsx file. Install openpyxl first:  pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [[cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    # CSV / TSV
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        delim = "\t" if sample.count("\t") > sample.count(",") else ","
        return [[cell_to_str(c) for c in row] for row in csv.reader(f, delimiter=delim)]


# ── Header / column detection ────────────────────────────────────────────────
def find_header(grid):
    for i, row in enumerate(grid):
        cells = [norm(c) for c in row]
        if "attribute" in cells and any("attribute_akeneo" == c or "akeneo" in c for c in cells):
            return i
    # Fallback: first non-empty row.
    for i, row in enumerate(grid):
        if any(str(c).strip() for c in row):
            return i
    return -1


def col_index(header, *predicates):
    for i, cell in enumerate(header):
        n = norm(cell)
        for p in predicates:
            if p(n):
                return i
    return -1


# ── Value inference ──────────────────────────────────────────────────────────
def to_number(v):
    try:
        return float(str(v).replace(",", ".").strip())
    except ValueError:
        return None


def infer_type(values):
    """Return (suggested_data_type_label, options_or_range, note)."""
    vals = [v for v in (x.strip() for x in values) if v]
    if not vals:
        return "Text", "", ""
    distinct = list(dict.fromkeys(vals))  # preserve order, unique
    low = [v.lower() for v in vals]

    # Boolean: strictly yes/no/true/false (avoid pure 0/1 which reads as numeric).
    if all(v in (BOOL_TRUE | BOOL_FALSE) for v in low) and any(v in {"yes", "no", "true", "false"} for v in low):
        return "Boolean (Yes/No)", "", ""

    nums = [to_number(v) for v in vals]
    if all(n is not None for n in nums):
        lo, hi = min(nums), max(nums)
        rng = f"Observed range: {fmt_num(lo)}–{fmt_num(hi)}"
        if all(float(n).is_integer() for n in nums):
            return "Number (integer)", rng, ""
        return "Number", rng, ""

    # Enum: a small set of short, repeated categorical strings.
    if len(distinct) <= 25 and max(len(v) for v in distinct) <= 60:
        note = "Only one option observed — verify" if len(distinct) == 1 else ""
        return "Simple select", "; ".join(sorted(distinct, key=str.lower)), note

    return "Text", "", ""


def fmt_num(n):
    return str(int(n)) if float(n).is_integer() else ("%g" % n)


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Prepare OriginFlow attribute + SKU import files.")
    ap.add_argument("input", help="Path to the source .xlsx or .csv")
    ap.add_argument("--category", required=True, help="Category name (used for output filenames)")
    ap.add_argument("--out-dir", default=".", help="Output directory (default: current dir)")
    ap.add_argument("--sku-prefix", default="100", help="SKU number prefix (default: 100)")
    args = ap.parse_args()

    grid = load_grid(args.input)
    hi = find_header(grid)
    if hi < 0:
        sys.exit("Could not find a header row containing 'Attribute' and 'attribute_akeneo'.")
    header = grid[hi]

    c_attr = col_index(header, lambda n: n == "attribute")
    c_code = col_index(header, lambda n: n == "attribute_akeneo", lambda n: n == "akeneo code", lambda n: n == "akeneo")
    c_where = col_index(header, lambda n: n == "where to store", lambda n: "where to store" in n)
    c_group = col_index(header, lambda n: n in ("group", "section", "family"))

    if c_attr < 0 or c_code < 0:
        sys.exit("Input must have an 'Attribute' column and an 'attribute_akeneo' column.")

    sku_re = re.compile(r"^\s*" + re.escape(args.sku_prefix) + r"\d+\s*$")
    sku_cols = [i for i, cell in enumerate(header) if sku_re.match(str(cell))]
    if not sku_cols:
        sys.exit(f"No SKU columns found (headers starting with '{args.sku_prefix}').")
    sku_numbers = [str(header[i]).strip() for i in sku_cols]

    warnings = []
    if c_where < 0:
        warnings.append("No 'Where to store' column found — keeping ALL attribute rows.")

    # Detect a title row (used to fill SKU titles); not emitted as an attribute.
    title_by_sku = {i: "" for i in sku_cols}

    attr_out = []          # rows for the attributes CSV
    sku_body = []          # [code, val_per_sku...] rows for the SKU CSV
    seen_codes = set()
    kept = 0
    type_counts = {}

    for r in grid[hi + 1:]:
        if not any(str(c).strip() for c in r):
            continue
        name = r[c_attr].strip() if c_attr < len(r) else ""
        if not name:
            continue

        # Title row -> fill SKU titles, then skip.
        if norm(name) in TITLE_ROW_NAMES:
            for i in sku_cols:
                if i < len(r) and r[i].strip():
                    title_by_sku[i] = r[i].strip()
            continue

        # Only keep rows flagged for Akeneo.
        if c_where >= 0:
            where = r[c_where].strip() if c_where < len(r) else ""
            if "akeneo" not in where.lower():
                continue

        code = (r[c_code].strip() if c_code < len(r) else "")
        note_parts = []
        if not code:
            code = slugify(name)
            note_parts.append("Akeneo code derived from name — verify")
        if code.lower() in seen_codes:
            note_parts.append("Duplicate Akeneo code in file")
        seen_codes.add(code.lower())

        values = [r[i].strip() if i < len(r) else "" for i in sku_cols]
        dtype, options, tnote = infer_type(values)
        if tnote:
            note_parts.append(tnote)
        type_counts[dtype] = type_counts.get(dtype, 0) + 1

        group = DEFAULT_GROUP
        if c_group >= 0 and c_group < len(r) and r[c_group].strip():
            group = GROUP_ALIASES.get(norm(r[c_group]), DEFAULT_GROUP)

        attr_out.append([name, group, code, dtype, options, "; ".join(note_parts)])
        sku_body.append([code] + values)
        kept += 1

    if kept == 0:
        sys.exit("No attribute rows kept. Check the 'Where to store' column contains 'Akeneo'.")

    # Duplicate / malformed SKU checks.
    dupes = {s for s in sku_numbers if sku_numbers.count(s) > 1}
    if dupes:
        warnings.append(f"Duplicate SKU columns: {', '.join(sorted(dupes))}")

    slug = slugify(args.category) or "category"
    os.makedirs(args.out_dir, exist_ok=True)
    attr_path = os.path.join(args.out_dir, f"{slug}_attributes.csv")
    sku_path = os.path.join(args.out_dir, f"{slug}_skus.csv")

    with open(attr_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(ATTR_HEADER)
        w.writerows(attr_out)

    with open(sku_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Attribute"] + sku_numbers)
        w.writerow(["Title"] + [title_by_sku[i] for i in sku_cols])
        w.writerows(sku_body)

    # ── Summary ──────────────────────────────────────────────────────────────
    print(f"\n✓ Wrote {attr_path}")
    print(f"✓ Wrote {sku_path}\n")
    print(f"Attributes kept: {kept}   SKUs: {len(sku_numbers)}")
    print("Types: " + ", ".join(f"{k}: {v}" for k, v in sorted(type_counts.items())))
    flagged = [a for a in attr_out if a[5]]
    if flagged:
        print(f"\nFlagged attributes ({len(flagged)}):")
        for a in flagged:
            print(f"  - {a[0]} [{a[2]}]: {a[5]}")
    if warnings:
        print("\nWarnings:")
        for wn in warnings:
            print(f"  - {wn}")
    print("\nNext: import the *_attributes.csv first (Admin > Categories), then the *_skus.csv "
          "(SKU Catalog > Bulk upload).")


if __name__ == "__main__":
    main()
