"""Turn the EE marking guide workbook into an idempotent seed SQL script.

Reads docs/EE marking guide_Hobs Induction hobs_R1_202507.xlsx and writes
db_migrations/seed_regulations_hobs_marking_guide.sql.

Sheet shape (verified, not assumed):
  R2  C8..C13  the source standard/regulation each requirement belongs to
  R3  C1..C6   headers: Class | Symbol Example | Rating label | Sales packaging | IM | (Product)
  R5..R48      one requirement per row; C1 = text, C3..C6 = per-surface marker,
               C8..C13 = clause reference in the column of its source
  R51..R53     legend: Wingdings 'u' = mandatory, 'û' = not necessary, 'O' = optional
"""

import io
import re
import openpyxl

SRC = 'docs/EE marking guide_Hobs Induction hobs_R1_202507.xlsx'
OUT = 'db_migrations/seed_regulations_hobs_marking_guide.sql'

# Marker glyphs are Wingdings; the legend in R51..R53 defines them.
MANDATORY, NOT_NEEDED, OPTIONAL = 'ü', 'û', 'O'
SURFACES = [(3, 'Rating label'), (4, 'Sales packaging'), (5, 'IM'), (6, 'Product')]

# One entry per regulation the sheet cites. `column` is its column in R2; `clause_match`
# splits the single column that names TWO regulations (Annex II & III = standby ecodesign,
# Annex I = the hob energy-information regulation). It is a predicate, not a prefix, because
# 'ANNEX II & III' starts with 'ANNEX I' -- a prefix test silently gave 66/2014 the standby
# rows too.
#
# reference_code values deliberately match the codes ALREADY in the library, so this
# script updates those rows instead of being refused by uq_regulations_reference_code.
REGULATIONS = [
    dict(column=8, code='EN IEC 60335-1:2021', jurisdiction='EU',
         title='Household and similar electrical appliances — Safety — Part 1: General requirements'),
    dict(column=9, code='EN IEC 60335-2-6:2024', jurisdiction='EU',
         title='Household and similar electrical appliances — Safety — Part 2-6: Particular '
               'requirements for stationary cooking ranges, hobs, ovens and similar appliances'),
    dict(column=10, code='Blue Guide 2022', jurisdiction='EU',
         title="The 'Blue Guide' on the implementation of EU product rules 2022"),
    dict(column=11, code='Directive 2012/19/EU WEEE', jurisdiction='EU',
         title='Directive 2012/19/EU on waste electrical and electronic equipment (WEEE)'),
    dict(column=12, code='UKCA marking guidance', jurisdiction='UK',
         title='Guidance on using the UKCA marking'),
    dict(column=13, clause_match=lambda c: c.upper().startswith('ANNEX II'),
         code='Regulation (EU) 2023/826', jurisdiction='EU',
         title='Commission Regulation (EU) 2023/826 — ecodesign requirements for off mode, '
               'standby mode and networked standby energy consumption'),
    dict(column=13, clause_match=lambda c: c.upper().strip() == 'ANNEX I',
         code='Regulation (EU) No 66/2014', jurisdiction='EU',
         title='Commission Regulation (EU) No 66/2014 — ecodesign requirements for domestic '
               'ovens, hobs and range hoods'),
]

PROVENANCE = ('Marking, labelling and instruction obligations for hobs / induction hobs, '
              'taken from the EE marking guide (Hobs / Induction hobs, R1, 2025-07).')


def flatten(text):
    """One requirement per line in the target column, so embedded newlines must go."""
    return re.sub(r'\s+', ' ', text.replace('\n', ' ')).strip()


def surfaces_for(ws, row):
    """(mandatory, optional) surface names for one requirement."""
    mandatory, optional = [], []
    for col, label in SURFACES:
        mark = str(ws.cell(row=row, column=col).value or '').strip()
        if mark == MANDATORY:
            mandatory.append(label)
        elif mark == OPTIONAL:
            optional.append(label)
    return mandatory, optional


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb['Marking_RL_SP_IM_Product']
    cell = lambda r, c: str(ws.cell(row=r, column=c).value or '').strip()

    # Collect requirement rows per source column.
    rows_by_column = {}
    for r in range(5, 49):
        text = cell(r, 1)
        if not text:
            continue
        source = next(((c, cell(r, c)) for c in range(8, 14) if cell(r, c)), None)
        if not source:
            print(f'  ! row {r} has no source column, skipped')
            continue
        column, clause = source
        rows_by_column.setdefault(column, []).append((r, clause, text))

    out = [
        '-- Seed: regulations from the EE marking guide (Hobs / Induction hobs, R1 2025-07).',
        '--',
        '-- GENERATED from docs/EE marking guide_Hobs Induction hobs_R1_202507.xlsx.',
        '-- Regenerate rather than hand-edit: the generator is .gen_reg_seed.py in the repo root',
        '-- history, and hand edits here are lost the next time the guide is revised.',
        '--',
        '-- NOT a schema migration -- data only, and safe to re-run. Deliberately named',
        '-- seed_* rather than 121_* so it is not mistaken for one.',
        '--',
        '-- WHAT IT WRITES, AND WHAT IT LEAVES ALONE',
        '--',
        '--   checklist, notes, title, jurisdiction -- OVERWRITTEN from the workbook. The guide',
        '--                 is the authoritative source, so re-running re-syncs all four and any',
        '--                 value edited in the app in the meantime is replaced.',
        '--   summary_md -- NOT WRITTEN, because this workbook contains none. It is a marking',
        '--                 GUIDE, not regulation text. Setting it from here would mean writing',
        '--                 NULL over the summaries already uploaded, and the AI check reads',
        '--                 summary_md and nothing else -- every future check would refuse to run.',
        '--   status     -- set on INSERT only. Not overwritten, so a regulation deliberately',
        '--                 retired as ''superseded'' is not silently brought back to active.',
        '--',
        '-- Five of these reference codes already exist in the library, so this is written as an',
        '-- upsert on the unique lower(btrim(reference_code)) index rather than plain inserts,',
        '-- which uq_regulations_reference_code would refuse.',
        '--',
        '-- REQUIRES migrations 116, 118, 119 and 120 to be applied first. 119 is the hard',
        '-- dependency (it adds regulations.checklist) and the preflight below checks for it.',
        '--',
        '-- The category association at the end is what makes the checklist appear on the induction',
        '-- hob templates (an active regulation listing a category applies to that category\'s',
        '-- templates -- migration 116). Delete that block if you would rather assign by hand.',
        '',
        'BEGIN;',
        '',
        '-- Preflight: this seed writes regulations.checklist, which migration 119 adds. Without',
        '-- that column the first INSERT fails with a bare "column does not exist"; say why instead.',
        'DO $chk$ BEGIN',
        '  IF NOT EXISTS (',
        '    SELECT 1 FROM information_schema.columns',
        "     WHERE table_schema = 'public' AND table_name = 'regulations' AND column_name = 'checklist'",
        '  ) THEN',
        "    RAISE EXCEPTION 'regulations.checklist is missing -- apply db_migrations/119_regulation_checklist.sql before this seed.';",
        '  END IF;',
        'END $chk$;',
        '',
    ]

    total_items = 0
    for reg in REGULATIONS:
        rows = rows_by_column.get(reg['column'], [])
        if 'clause_match' in reg:
            rows = [x for x in rows if reg['clause_match'](x[1])]
        if not rows:
            print(f"  ! no rows for {reg['code']}, skipped")
            continue

        items = []
        for _row, clause, text in rows:
            mandatory, optional = surfaces_for(ws, _row)
            where = ', '.join(mandatory) if mandatory else 'not required on any surface'
            if optional:
                where += f" (optional: {', '.join(optional)})"
            items.append(f'{clause} · {where} — {flatten(text)}')
        total_items += len(items)

        checklist = '\n'.join(items)
        notes = f'{PROVENANCE} Clause references are to {reg["code"]}.'

        out += [
            f'-- {reg["code"]} — {len(items)} checklist item(s)',
            'INSERT INTO public.regulations (title, reference_code, jurisdiction, notes, checklist, status)',
            'VALUES (',
            f'  $sql${reg["title"]}$sql$,',
            f'  $sql${reg["code"]}$sql$,',
            f'  $sql${reg["jurisdiction"]}$sql$,',
            f'  $sql${notes}$sql$,',
            f'  $sql${checklist}$sql$,',
            "  'active'",
            ')',
            'ON CONFLICT (lower(btrim(reference_code))) DO UPDATE SET',
            '  checklist    = EXCLUDED.checklist,',
            '  notes        = EXCLUDED.notes,',
            '  title        = EXCLUDED.title,',
            '  jurisdiction = EXCLUDED.jurisdiction,',
            '  updated_at   = NOW();',
            '',
        ]

    out += [
        '-- Associate all seven with the induction hob category, which is what surfaces the',
        '-- checklist on its templates (an active regulation listing a category applies to that',
        '-- category\'s templates -- migration 116). The name match is deliberately loose because',
        '-- the category may be named "Induction hob", "Induction hobs" or "Hobs / Induction hobs".',
        '--',
        '-- Written as a set union rather than `UPDATE ... FROM categories_l3` + array_append: that',
        '-- join form appends only ONE category id when several names match, because the target row',
        '-- is updated once against an arbitrarily chosen joined row. The union below adds every',
        '-- match, keeps categories already present (so other categories are never stripped), and',
        '-- is idempotent on a re-run.',
        'UPDATE public.regulations r',
        '   SET applicable_categories = COALESCE((',
        '         SELECT array_agg(DISTINCT x)',
        '           FROM (',
        '                 SELECT unnest(r.applicable_categories) AS x',
        '                  UNION',
        "                 SELECT c.id::text FROM public.categories_l3 c WHERE c.name ILIKE '%induction%'",
        '                ) s',
        "       ), '{}'),",
        '       updated_at = NOW()',
        ' WHERE lower(btrim(r.reference_code)) IN (',
        ',\n'.join(f"         lower(btrim($sql${reg['code']}$sql$))" for reg in REGULATIONS),
        ' );',
        '',
        '-- Sanity check: what the script just wrote. Nothing is committed if this looks wrong --',
        '-- ROLLBACK instead of COMMIT.',
        'SELECT reference_code,',
        '       jurisdiction,',
        "       array_length(string_to_array(checklist, E'\\n'), 1) AS checklist_items,",
        '       (summary_md IS NOT NULL) AS has_summary,',
        '       applicable_categories',
        '  FROM public.regulations',
        ' WHERE lower(btrim(reference_code)) IN (',
        ',\n'.join(f"         lower(btrim($sql${reg['code']}$sql$))" for reg in REGULATIONS),
        ' )',
        ' ORDER BY reference_code;',
        '',
        'COMMIT;',
        '',
        "NOTIFY pgrst, 'reload schema';",
        '',
    ]

    io.open(OUT, 'w', encoding='utf-8', newline='\n').write('\n'.join(out))
    print(f'wrote {OUT}: {len(REGULATIONS)} regulations, {total_items} checklist items')


main()
